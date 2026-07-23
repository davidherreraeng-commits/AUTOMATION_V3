from pathlib import Path

import pytest
from openpyxl import Workbook

from adapters.input.excel import (
    ExcelContractSource,
    ExcelImportError,
    MissingRequiredColumnsError,
)


VALID_HEADERS = [
    "No. de Contrato",
    "Cédula o Nit Contratista",
    "Tipo Persona",
    "Código del Proyecto",
    "Objeto del Contrato",
    "Fecha de Suscripción",
    "Fecha de Inicio",
    "Valor",
    "Plazo Estimado (En Dias)",
    "Modalidad o Proceso",
    "Procedimiento/Causal",
    "Tipo de Contrato",
    "Rubro Presupuestal",
    "Sub-Sector",
    "Cédula Supervisor",
    "No. CDP",
]


VALID_ROW = [
    "70-2026",
    "900469775-8",
    "Jurídica",
    "I-23021-2026",
    "Servicio de software institucional.",
    "20/01/2026",
    "21/01/2026",
    "$ 1.476.190",
    180,
    "Contratación Directa",
    "Prestación de Servicios",
    "Servicios",
    "IDEA-2026 - RECURSOS CONVENIO IDEA",
    "Tecnología",
    71693738,
    235097,
]


def create_workbook(
    file_path: Path,
    *,
    sheet_name: str = "Contratos",
    headers=None,
    rows=None,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name

    worksheet.append(
        headers if headers is not None else VALID_HEADERS
    )

    for row in rows or []:
        worksheet.append(row)

    workbook.save(file_path)
    workbook.close()


def create_source(
    file_path: Path,
    *,
    sheet_name: str | None = None,
) -> ExcelContractSource:
    return ExcelContractSource(
        file_path=file_path,
        default_dependency="Proyectos Especiales",
        default_budget_year=2026,
        sheet_name=sheet_name,
    )


def test_should_read_valid_contract(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "contratos.xlsx"

    create_workbook(
        file_path,
        rows=[VALID_ROW],
    )

    results = list(
        create_source(file_path).read()
    )

    assert len(results) == 1
    assert results[0].is_valid
    assert results[0].row_number == 2
    assert results[0].contract is not None
    assert (
        results[0].contract.contract_number
        == "70-2026"
    )


def test_should_return_invalid_result_and_continue(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "contratos.xlsx"

    invalid_row = list(VALID_ROW)
    invalid_row[0] = None

    second_valid_row = list(VALID_ROW)
    second_valid_row[0] = "71-2026"

    create_workbook(
        file_path,
        rows=[
            invalid_row,
            second_valid_row,
        ],
    )

    results = list(
        create_source(file_path).read()
    )

    assert len(results) == 2

    assert results[0].is_invalid
    assert results[0].row_number == 2

    assert results[1].is_valid
    assert results[1].row_number == 3
    assert results[1].contract is not None
    assert (
        results[1].contract.contract_number
        == "71-2026"
    )


def test_should_ignore_completely_empty_rows(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "contratos.xlsx"

    create_workbook(
        file_path,
        rows=[
            [None] * len(VALID_HEADERS),
            VALID_ROW,
        ],
    )

    results = list(
        create_source(file_path).read()
    )

    assert len(results) == 1
    assert results[0].row_number == 3


def test_should_reject_missing_required_headers(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "contratos.xlsx"

    headers = [
        header
        for header in VALID_HEADERS
        if header != "Código del Proyecto"
    ]

    create_workbook(
        file_path,
        headers=headers,
        rows=[],
    )

    with pytest.raises(
        MissingRequiredColumnsError,
    ):
        list(
            create_source(file_path).read()
        )


def test_should_read_selected_sheet(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "contratos.xlsx"

    workbook = Workbook()

    first_sheet = workbook.active
    first_sheet.title = "Resumen"
    first_sheet.append(["Información"])
    first_sheet.append(["No procesar"])

    contract_sheet = workbook.create_sheet(
        "Contratos"
    )
    contract_sheet.append(VALID_HEADERS)
    contract_sheet.append(VALID_ROW)

    workbook.save(file_path)
    workbook.close()

    results = list(
        create_source(
            file_path,
            sheet_name="Contratos",
        ).read()
    )

    assert len(results) == 1
    assert results[0].is_valid


def test_should_reject_unknown_sheet(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "contratos.xlsx"

    create_workbook(
        file_path,
        rows=[VALID_ROW],
    )

    with pytest.raises(
        ExcelImportError,
        match="no existe",
    ):
        list(
            create_source(
                file_path,
                sheet_name="Hoja inexistente",
            ).read()
        )