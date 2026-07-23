from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from application.dto import ContractImportResult
from application.ports import ContractSource

from adapters.input.excel.column_mapper import (
    ColumnMapper,
    ColumnMapping,
)
from adapters.input.excel.errors import ExcelImportError
from adapters.input.excel.row_to_contract_mapper import (
    ContractRowMapper,
)
from adapters.input.excel.value_normalizer import ValueNormalizer


class ExcelContractSource(ContractSource):
    """
    Fuente de contratos basada en archivos Excel.

    El archivo se abre en modo lectura y `data_only=True`, por lo que
    las fórmulas se leen usando su último valor calculado almacenado.

    Los errores de una fila producen un ContractImportResult inválido.
    Los errores estructurales del archivo lanzan ExcelImportError o
    una subclase de HeaderMappingError.
    """

    SUPPORTED_EXTENSIONS: frozenset[str] = frozenset(
        {
            ".xlsx",
            ".xlsm",
        }
    )

    def __init__(
        self,
        file_path: str | Path,
        *,
        default_dependency: str | None,
        default_budget_year: int,
        sheet_name: str | None = None,
        header_row: int = 1,
        column_mapper: ColumnMapper | None = None,
        row_mapper: ContractRowMapper | None = None,
    ) -> None:
        self._file_path = Path(file_path)
        self._sheet_name = sheet_name
        self._header_row = header_row

        if self._header_row < 1:
            raise ValueError(
                "La fila de encabezados debe ser igual o superior a 1."
            )

        self._column_mapper = (
            column_mapper
            if column_mapper is not None
            else ColumnMapper()
        )

        self._row_mapper = (
            row_mapper
            if row_mapper is not None
            else ContractRowMapper(
                default_dependency=default_dependency,
                default_budget_year=default_budget_year,
            )
        )

    @property
    def file_path(self) -> Path:
        return self._file_path

    @property
    def sheet_name(self) -> str | None:
        return self._sheet_name

    def read(self) -> Iterable[ContractImportResult]:
        """
        Lee las filas del archivo de manera incremental.

        El workbook permanece abierto durante la iteración y se cierra
        automáticamente al terminar o al producirse una excepción.
        """

        self._validate_file()

        try:
            workbook = load_workbook(
                filename=self._file_path,
                read_only=True,
                data_only=True,
            )
        except (
            InvalidFileException,
            OSError,
            ValueError,
        ) as error:
            raise ExcelImportError(
                f"No se pudo abrir el archivo Excel "
                f"'{self._file_path}': {error}"
            ) from error

        try:
            worksheet = self._resolve_worksheet(
                workbook
            )

            headers = self._read_headers(
                worksheet
            )

            column_mapping = (
                self._column_mapper.map_headers(
                    headers
                )
            )

            yield from self._read_rows(
                worksheet=worksheet,
                headers=headers,
                column_mapping=column_mapping,
            )

        finally:
            workbook.close()

    def _validate_file(self) -> None:
        if not self._file_path.exists():
            raise ExcelImportError(
                f"El archivo no existe: {self._file_path}"
            )

        if not self._file_path.is_file():
            raise ExcelImportError(
                f"La ruta no corresponde a un archivo: "
                f"{self._file_path}"
            )

        suffix = self._file_path.suffix.lower()

        if suffix not in self.SUPPORTED_EXTENSIONS:
            raise ExcelImportError(
                "Formato de archivo no soportado. "
                f"Extensión recibida: '{suffix}'. "
                "Use .xlsx o .xlsm."
            )

    def _resolve_worksheet(
        self,
        workbook,
    ) -> Worksheet:
        if self._sheet_name is None:
            return workbook.active

        if self._sheet_name not in workbook.sheetnames:
            raise ExcelImportError(
                f"La hoja '{self._sheet_name}' no existe. "
                "Hojas disponibles: "
                + ", ".join(workbook.sheetnames)
            )

        return workbook[self._sheet_name]

    def _read_headers(
        self,
        worksheet: Worksheet,
    ) -> tuple[Any, ...]:
        row_iterator = worksheet.iter_rows(
            min_row=self._header_row,
            max_row=self._header_row,
            values_only=True,
        )

        headers = next(
            row_iterator,
            None,
        )

        if headers is None:
            raise ExcelImportError(
                "No fue posible leer la fila de encabezados."
            )

        if all(
            ValueNormalizer.is_missing(header)
            for header in headers
        ):
            raise ExcelImportError(
                f"La fila {self._header_row} no contiene encabezados."
            )

        return tuple(headers)

    def _read_rows(
        self,
        *,
        worksheet: Worksheet,
        headers: tuple[Any, ...],
        column_mapping: ColumnMapping,
    ) -> Iterable[ContractImportResult]:
        source_headers = tuple(
            self._normalize_source_header(header)
            for header in headers
        )

        for row_number, row_values in enumerate(
            worksheet.iter_rows(
                min_row=self._header_row + 1,
                values_only=True,
            ),
            start=self._header_row + 1,
        ):
            if self._is_empty_row(row_values):
                continue

            source_row = self._build_source_row(
                source_headers=source_headers,
                row_values=row_values,
            )

            canonical_row = (
                column_mapping.canonicalize_row(
                    source_row
                )
            )

            yield self._row_mapper.map(
                row_number=row_number,
                canonical_row=canonical_row,
            )

    @staticmethod
    def _normalize_source_header(
        header: Any,
    ) -> str | None:
        if header is None:
            return None

        normalized = str(header).strip()

        return normalized or None

    @staticmethod
    def _is_empty_row(
        row_values: tuple[Any, ...],
    ) -> bool:
        return all(
            ValueNormalizer.is_missing(value)
            for value in row_values
        )

    @staticmethod
    def _build_source_row(
        *,
        source_headers: tuple[str | None, ...],
        row_values: tuple[Any, ...],
    ) -> dict[str, Any]:
        source_row: dict[str, Any] = {}

        for index, source_header in enumerate(
            source_headers
        ):
            if source_header is None:
                continue

            value = (
                row_values[index]
                if index < len(row_values)
                else None
            )

            source_row[source_header] = value

        return source_row